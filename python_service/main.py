import io
import json
import logging
import os
import uuid
import shutil
import tempfile
import time
from datetime import datetime
from typing import List, Dict, Any, Optional

import pandas as pd
# import whisper  <-- Removed
from faster_whisper import WhisperModel # <-- Added
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks, Depends
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import range_boundaries
from sqlalchemy import create_engine, Column, String, DateTime, JSON, Text
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session

from fastapi.staticfiles import StaticFiles

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("yz-portal-backend")

# 确保上传目录存在
UPLOAD_DIR = "uploads"
if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)

# ==========================================
# Database Setup
# ==========================================
DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://postgres:postgres@localhost:5432/yz-portal")

# Retry connection logic (db might not be ready immediately)
engine = None
SessionLocal = None
Base = declarative_base()

class MediaTask(Base):
    __tablename__ = "media_tasks"

    id = Column(String, primary_key=True, index=True)
    file_name = Column(String, nullable=False)
    task_type = Column(String, default="transcribe") # transcribe, translate
    status = Column(String, default="pending") # pending, processing, completed, failed
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    result = Column(JSON, nullable=True)
    error_message = Column(Text, nullable=True)
    
    # Extra fields to match frontend expectations roughly
    start_time = Column(DateTime, default=datetime.utcnow) # alias for created_at logic

def init_db():
    global engine, SessionLocal
    try:
        engine = create_engine(DATABASE_URL)
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
        Base.metadata.create_all(bind=engine)
        logger.info("Database initialized successfully.")
    except Exception as e:
        logger.error(f"Database initialization failed: {e}")

# Dependency
def get_db():
    if SessionLocal is None:
        init_db()
    if SessionLocal is None:
         raise HTTPException(status_code=500, detail="Database not initialized")
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ==========================================
# Whisper Model Setup
# ==========================================
model = None

def load_whisper_model():
    global model
    if model is None:
        try:
            logger.info("Loading Faster Whisper model (base)...")
            # compute_type="int8" is much faster on CPU
            model = WhisperModel("base", device="cpu", compute_type="int8")
            logger.info("Faster Whisper model loaded.")
        except Exception as e:
            logger.error(f"Failed to load Whisper model: {e}")

# ==========================================
# FastAPI App
# ==========================================
app = FastAPI(title="YZ Portal Backend Service")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 挂载静态文件目录
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

@app.on_event("startup")
def startup_event():
    init_db()
    # Lazy load whisper to allow fast startup, or load here if preferred
    # load_whisper_model() 

# ==========================================
# Excel Parser Logic (Legacy)
# ==========================================
class ExcelParser:
    """
    高级 Excel 解析引擎
    """
    def __init__(self, file_content: bytes, header_rows: int = 1):
        self.file_content = file_content
        self.header_rows = header_rows
        self.wb = None

    def process(self) -> Dict[str, Any]:
        try:
            self.wb = load_workbook(io.BytesIO(self.file_content), data_only=True)
        except Exception as e:
            logger.error(f"Loading failed: {e}")
            raise HTTPException(status_code=400, detail="Invalid Excel file.")

        result_sheets = {}
        total_rows = 0
        total_cols = 0

        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            self._semantic_fill_merged_cells(ws)
            data_blocks = self._detect_data_blocks(ws)

            sheet_md_output = []
            sheet_json_output = []

            for idx, block_range in enumerate(data_blocks):
                df_raw = self._block_to_dataframe(ws, block_range)
                if df_raw is None or df_raw.empty: continue

                df_json = self._flatten_headers(df_raw.copy())
                df_json = self._clean_content(df_json, is_md=False)
                sheet_json_output.extend(df_json.to_dict(orient="records"))

                df_md = df_raw.copy()
                df_md = self._clean_content(df_md, is_md=True)
                
                if not df_md.empty:
                    md_header = df_md.iloc[0].tolist()
                    md_body = df_md.iloc[1:]
                    table_str = md_body.to_markdown(headers=md_header, index=False, tablefmt="github")
                    table_title = f"### {sheet_name} - Table {idx + 1}"
                    sheet_md_output.append(f"{table_title}\n\n{table_str}")

                total_rows += len(df_json)
                total_cols = max(total_cols, len(df_json.columns))

            result_sheets[sheet_name] = {
                "md": "\n\n".join(sheet_md_output),
                "json": sheet_json_output
            }

        final_md = "\n\n".join([f"## Sheet: {k}\n{v['md']}" for k, v in result_sheets.items()])
        final_json = {k: v['json'] for k, v in result_sheets.items()}

        return {
            "rows": total_rows,
            "columns": total_cols,
            "data": {
                "md_format": final_md,
                "json_format": json.dumps(final_json, ensure_ascii=False, default=str)
            }
        }

    def _semantic_fill_merged_cells(self, ws: Worksheet):
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            val = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merged_range))
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    ws.cell(row=r, column=c).value = val

    def _detect_data_blocks(self, ws: Worksheet) -> List[tuple]:
        valid_rows = []
        for row in ws.iter_rows():
            if any(cell.value is not None and str(cell.value).strip() != "" for cell in row):
                valid_rows.append(row[0].row)
        if not valid_rows: return []
        blocks = []
        start = valid_rows[0]
        prev = valid_rows[0]
        for curr in valid_rows[1:]:
            if curr - prev > 2: 
                blocks.append((start, prev))
                start = curr
            prev = curr
        blocks.append((start, prev))
        return blocks

    def _block_to_dataframe(self, ws: Worksheet, row_range: tuple) -> Optional[pd.DataFrame]:
        start, end = row_range
        data = []
        for row in ws.iter_rows(min_row=start, max_row=end, values_only=True):
            data.append(list(row))
        if not data: return None
        df = pd.DataFrame(data)
        return df.dropna(axis=0, how='all').dropna(axis=1, how='all')

    def _flatten_headers(self, df: pd.DataFrame) -> pd.DataFrame:
        h_rows = max(1, self.header_rows)
        if len(df) <= h_rows: return df
        header_df = df.iloc[:h_rows]
        data_df = df.iloc[h_rows:].reset_index(drop=True)
        new_cols = []
        for c in range(df.shape[1]):
            parts = [str(header_df.iloc[r, c]).strip() for r in range(h_rows) if pd.notna(header_df.iloc[r, c]) and str(header_df.iloc[r, c]).strip() != ""]
            name = " > ".join(parts) if parts else f"Col_{c+1}"
            new_cols.append(name)
        final_cols = []
        seen = {}
        for col in new_cols:
            if col in seen:
                seen[col] += 1
                final_cols.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                final_cols.append(col)
        data_df.columns = final_cols
        return data_df

    def _clean_content(self, df: pd.DataFrame, is_md: bool) -> pd.DataFrame:
        df = df.fillna("")
        def clean(v):
            if is_md and isinstance(v, str):
                v = v.strip().replace('\r\n', '\n').replace('\r', '\n')
                v = v.replace('\n', '<br/>')
                v = v.replace('|', '\|')
                return v
            return v
        return df.applymap(clean)

@app.post("/parse/excel")
async def parse_excel(file: UploadFile = File(...), header_rows: int = Form(1)):
    if not file.filename.endswith(('.xlsx', '.xls', '.xlsm')):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")
    try:
        content = await file.read()
        etl = ExcelParser(content, header_rows)
        result = etl.process()
        result["filename"] = file.filename
        return result
    except Exception as e:
        logger.error(f"Fail: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==========================================
# Media Transcription Logic
# ==========================================

def process_transcription(task_id: str, file_path: str, model_size: str = "base"):
    """
    Background task to run Whisper
    """
    logger.info(f"Starting transcription for task {task_id}")
    db = SessionLocal()
    task = db.query(MediaTask).filter(MediaTask.id == task_id).first()
    
    if not task:
        logger.error(f"Task {task_id} not found in background worker")
        if os.path.exists(file_path):
            os.remove(file_path)
        return

    try:
        task.status = "processing"
        db.commit()

        load_whisper_model() # Ensure model is loaded
        
        # Run Whisper
        # Note: 'model' global variable should be safe if we assume single worker or read-only model access
        # If model loading failed, this will raise error
        if model is None:
             raise Exception("Whisper model failed to load.")

        # faster-whisper returns (segments, info)
        segments_generator, info = model.transcribe(file_path, beam_size=5)
        
        # segments is a generator, convert to list
        segments_list = []
        full_text = []
        
        for segment in segments_generator:
            segments_list.append({
                "start": segment.start,
                "end": segment.end,
                "text": segment.text,
                "speaker": "Speaker 0" # Default speaker
            })
            full_text.append(segment.text)

        transcription_result = {
            "text": "".join(full_text),
            "language": info.language,
            "segments": segments_list
        }

        task.result = transcription_result
        task.status = "completed"
        logger.info(f"Task {task_id} completed")

    except Exception as e:
        logger.error(f"Task {task_id} failed: {e}")
        task.status = "failed"
        task.error_message = str(e)
    finally:
        task.updated_at = datetime.utcnow()
        db.commit()
        db.close()
        # 不再删除文件，因为前端需要播放它
        # if os.path.exists(file_path): ...

@app.post("/speech-to-text")
async def create_media_task(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    task: str = "transcribe", # Query param by default
    model: str = "base",      # Query param by default
    language: Optional[str] = None, # Query param by default
    db: Session = Depends(get_db)
):
    # 1. Generate ID
    task_id = str(uuid.uuid4())
    
    # 2. Save file to persistent uploads directory
    file_ext = os.path.splitext(file.filename)[1]
    save_filename = f"{task_id}{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, save_filename)
    
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save file: {e}")
        
    # 3. Create DB Entry
    new_task = MediaTask(
        id=task_id,
        file_name=file.filename,
        task_type=task,
        status="pending",
        start_time=datetime.utcnow()
    )
    db.add(new_task)
    db.commit()
    db.refresh(new_task)
    
    # 4. Trigger Background Task
    background_tasks.add_task(process_transcription, task_id, file_path, model)
    
    # 5. Return Response
    return {
        "code": 200,
        "msg": "success",
        "data": {
            "identifier": task_id, # Frontend expects "identifier"
            "message": "Task created"
        }
    }

@app.get("/task/all")
async def get_media_task_list(db: Session = Depends(get_db)):
    tasks = db.query(MediaTask).order_by(MediaTask.created_at.desc()).all()
    # Format to match MediaTaskList
    res = []
    for t in tasks:
        # Calculate duration if possible from result
        duration = 0
        if t.result and "segments" in t.result and t.result["segments"]:
             duration = t.result["segments"][-1].get("end", 0)

        res.append({
            "identifier": t.id, # Frontend uses "identifier"
            "file_name": t.file_name,
            "task_type": t.task_type,
            "status": t.status,
            "language": t.result.get("language", "unknown") if t.result else "unknown",
            "start_time": t.start_time.isoformat() if t.start_time else t.created_at.isoformat(),
            "duration": duration,
            "url": None, # URL not available yet
            "error": t.error_message
        })
    return {"tasks": res}

@app.get("/task/{task_id}")
async def get_media_task_detail(task_id: str, db: Session = Depends(get_db)):
    task = db.query(MediaTask).filter(MediaTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    # Calculate duration
    duration = 0
    if task.result and "segments" in task.result and task.result["segments"]:
            duration = task.result["segments"][-1].get("end", 0)

    # Construct response strictly matching MediaTaskDetail interface
    # URL is mapped via Nginx or direct access
    file_ext = os.path.splitext(task.file_name)[1]
    file_url = f"/py-api/uploads/{task.id}{file_ext}"

    return {
        "identifier": task.id, # Added identifier
        "status": task.status,
        "result": task.result, # This contains { segments: [...] }
        "metadata": {
            "language": task.result.get("language", "unknown") if task.result else "unknown",
            "start_time": task.start_time.isoformat() if task.start_time else task.created_at.isoformat(),
            "end_time": task.updated_at.isoformat(),
            "file_name": task.file_name,
            "dutation": duration, # typo in frontend interface "dutation"
            "url": file_url
        }
    }

@app.delete("/task/{task_id}/delete")
async def delete_media_task(task_id: str, db: Session = Depends(get_db)):
    task = db.query(MediaTask).filter(MediaTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    db.delete(task)
    db.commit()
    return {"code": 200, "msg": "success"}

class UpdateResultRequest(BaseModel):
    result: Dict[str, Any]

@app.post("/task/{task_id}/update_result")
async def update_media_task_result(task_id: str, request: UpdateResultRequest, db: Session = Depends(get_db)):
    task = db.query(MediaTask).filter(MediaTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    task.result = request.result
    task.updated_at = datetime.utcnow()
    db.commit()
    return {"code": 200, "msg": "success"}

@app.get("/")
async def root():
    return {"message": "Service is running"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=9000)
